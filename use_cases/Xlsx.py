from loguru import logger
import openpyxl # type: ignore
from use_cases.PrintAutomation import PrintAutomation

class Xlsx:
    def __init__(self, file_path: str):
        """Classe para manipulação de arquivos Excel (.xlsx) usando openpyxl.
        :param file_path: Caminho completo do arquivo Excel a ser manipulado."""
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.printautomation = PrintAutomation()

    def get_sheet_names(self):
        """Retorna uma lista com os nomes das planilhas no arquivo Excel."""
        return self.workbook.sheetnames
    
    def get_sheet(self, sheet_name: str):
        """Retorna a planilha especificada pelo nome.
        :param sheet_name: Nome da planilha a ser retornada.
        :return: Objeto da planilha."""
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")
        return self.workbook[sheet_name]
    
    def get_cell_value(self, sheet_name: str, cell_reference: str):
        """Retorna o valor da célula especificada.
        :param sheet_name: Nome da planilha que contém a célula.
        :param cell_reference: Referência da célula (ex: 'A1').
        :return: Valor da célula."""
        sheet = self.get_sheet(sheet_name)
        return sheet[cell_reference].value
    
    def set_cell_value(self, sheet_name: str, cell_reference: str, value):
        """Define o valor da célula especificada.
        :param sheet_name: Nome da planilha que contém a célula.
        :param cell_reference: Referência da célula (ex: 'A1').
        :param value: Valor a ser definido na célula."""
        sheet = self.get_sheet(sheet_name)
        sheet[cell_reference].value = value
        self._save_workbook()

    def _save_workbook(self):
        try:
            self.workbook.save(self.file_path)
        except Exception as error_x:
            logger.critical(f'Erro ao salvar o arquivo "{self.file_path}": {error_x}')
            self.printautomation.print_error()
            raise RuntimeError(f'Erro ao salvar o arquivo "{self.file_path}": {error_x}') from error_x


if __name__ == '__main__':
    # Exemplo de uso
    xlsx = Xlsx("C:\\Zallpy\\PythonAutomationTemplate\\workbooks\\testes\\arquivo_excel.xlsx")
    print(f"Workbook loaded from: {xlsx.file_path}")
    print(f"Sheet names: {xlsx.get_sheet_names()}")
    sheet = xlsx.get_sheet("Planilha1")
    print(f"Accessing sheet: {sheet.title}")
    cell_value = xlsx.get_cell_value("Planilha1", "A1")
    print(f"Value in cell A1: {cell_value}")
    xlsx.set_cell_value("Planilha1", "A1", "1")
    print("Cell value updated.")
